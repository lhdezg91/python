
#.txt

#text_file =open("Intermedio/myfile.txt", "w+") #w+ es para escribir y leer pero sobre escribe el archivo
#text_file =open("Intermedio/myfile.txt", "a") #a es para escribir y leer pero no sobre escribe el archivo
#text_file =open("Intermedio/myfile.txt", "r") #r es para leer el archivo
text_file =open("Intermedio/myfile.txt", "r+") #r+ es para leer y escribir el archivo


print(text_file.read())#lee todo el archivo
print(text_file.readline())#lee la primera linea
print(text_file.readline())#lee la segunda linea

text_file.seek(0)#mueve el cursor al inicio del archivo
print(text_file.readline())#lee la primera linea

for line in text_file:
    print(line)


text_file.write("\nI'm a new line")
text_file.seek(0)


#text_file.write("Hello, world!\n")
#text_file.write("How are you?\n")

import os
#os.remove("Intermedio/myfile.txt")#borra el archivo
#os.rename("Intermedio/myfile.txt", "Intermedio/myfile_renamed.txt")#renombra el archivo
#os.mkdir("Intermedio/mydir")#crea un directorio
#os.rmdir("Intermedio/mydir")#borra un directorio


text_file.close()

#.json

import json

data = {
    "name": "John",
    "age": 30
}

with open("Intermedio/data.json", "w") as write_file:
    json.dump(data, write_file)

with open("Intermedio/data.json", "r") as read_file:
    data = json.load(read_file)
    print(data)
    print(data["name"])
    print(data["age"])
    
#os.remove("Intermedio/data.json")#borra el archivo

#.csv

import csv

with open("Intermedio/data.csv", mode="w") as write_file:
    writer = csv.writer(write_file)
    writer.writerow(["name", "age"])
    writer.writerow(["John", 30])
    writer.writerow(["Sara", 25])

with open("Intermedio/data.csv", mode="r") as read_file:
    reader = csv.reader(read_file)
    for row in reader:
        print(row)
        
#os.remove("Intermedio/data.csv")#borra el archivo

#.xlsx

import openpyxl

workbook = openpyxl.Workbook()
workbook.save("Intermedio/data.xlsx")

workbook = openpyxl.load_workbook("Intermedio/data.xlsx")
sheet = workbook.active

sheet["A1"] = "name"
sheet["B1"] = "age"
sheet["A2"] = "John"
sheet["B2"] = 30
sheet["A3"] = "Sara"
sheet["B3"] = 25

workbook.save("Intermedio/data.xlsx")

workbook = openpyxl.load_workbook("Intermedio/data.xlsx")
sheet = workbook.active

print(sheet["A1"].value)
print(sheet["B1"].value)
print(sheet["A2"].value)
print(sheet["B2"].value)
print(sheet["A3"].value)
print(sheet["B3"].value)

#os.remove("Intermedio/data.xlsx")#borra el archivo

#.pdf

from fpdf import FPDF

pdf = FPDF()
pdf.add_page()
pdf.set_font("Arial", size=12)
pdf.cell(200, 10, txt="Welcome to Python!", ln=True, align="C")
pdf.output("Intermedio/data.pdf")

#os.remove("Intermedio/data.pdf")#borra el archivo

#.zip

import zipfile

with zipfile.ZipFile("Intermedio/data.zip", "w") as newzip:
    newzip.write("Intermedio/data.json")
    newzip.write("Intermedio/data.csv")
    newzip.write("Intermedio/data.xlsx")
    newzip.write("Intermedio/data.pdf")
    
with zipfile.ZipFile("Intermedio/data.zip", "r") as newzip:
    newzip.extractall("Intermedio/data")
    
#os.remove("Intermedio/data.zip")#borra el archivo
#os.remove("Intermedio/data.json")#borra el archivo
#os.remove("Intermedio/data.csv")#borra el archivo
#os.remove("Intermedio/data.xlsx")#borra el archivo
#os.remove("Intermedio/data.pdf")#borra el archivo
#os.rmdir("Intermedio/data")#borra el directorio

#.tar

import tarfile

with tarfile.open("Intermedio/data.tar", "w") as newtar:
    newtar.add("Intermedio/data.json")
    newtar.add("Intermedio/data.csv")
    newtar.add("Intermedio/data.xlsx")
    newtar.add("Intermedio/data.pdf")
    
with tarfile.open("Intermedio/data.tar", "r") as newtar:
    newtar.extractall("Intermedio/data")
    
#os.remove("Intermedio/data.tar")#borra el archivo
#os.remove("Intermedio/data.json")#borra el archivo
#os.remove("Intermedio/data.csv")#borra el archivo
#os.remove("Intermedio/data.xlsx")#borra el archivo
#os.remove("Intermedio/data.pdf")#borra el archivo

